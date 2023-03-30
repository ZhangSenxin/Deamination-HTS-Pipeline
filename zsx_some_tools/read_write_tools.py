# -*- coding: UTF-8 -*-
# @Author: Zhang Senxin

"""
Functions face io or process on files
some functions used as linux
"""

import os
from collections.abc import Iterable
import pandas as pd
import numpy as np
import json

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    openpyxl_exist = True
except ModuleNotFoundError:
    openpyxl_exist = False
    print('\'openpyxl\' Module Not Found. Excel related function cannot to call.')

# 脚本路径
import sys
import inspect, sys

_code_path = sys.path[0].replace('\\', '/') + '/'
# _code_name = inspect.getsourcefile(sys._getframe(1))

# 用于追加输出logging信息，可以在某环节进行覆盖本行的输出
def my_logging(file, string, adding=False):
    """
    Support to write the logging info in the form like end='\r'
    :param file: the open file object
    :param string: string object to write
    :param adding: if True, just like '\a' form; if False, '\r'
    :return: None
    """
    if adding:
        file.seek(0, 2)
    string1 = (string + '\n').encode()
    file.write(string1)
    pos = file.tell()
    file.seek(pos - len(string1), 0)


def mkdir(path, file=False, silence=False):
    """
    Function to make direct direction.
    :param path: the direct direction
    :param file: itegrate to perform my_logging(file, path + '  - Successfully create', adding=True)
    :param silence: if print or not
    :return: None
    """
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path = path.rstrip("/")
    # 判断路径是否存在
    isExists = os.path.exists(path)
    if not isExists:
        # 如果不存在则创建目录
        os.makedirs(path)
        string = path + '  - Successfully create'
        if file:
            my_logging(file, string, adding=True)
        if not silence:
            print(string)

    else:
        # 如果目录存在则不创建，并提示目录已存在
        string = path + '  - Directory already exists'
        if file:
            my_logging(file, string, adding=True)
        if not silence:
            print(string)


# wc
def wc_py(path, time_print=False):
    """
    To get file line number as linux wc -l
    :param path: file path
    :param time_print: if print time used
    :return: line number
    """
    if time_print:
        import time
        start = time.time()

    with open(path, 'rb') as f:
        count = 0
        last_data = '\n'
        while True:
            data = f.read(0x400000)
            if not data:
                break
            count += data.count(b'\n')
            last_data = data
        if last_data[-1:] != b'\n':
            count += 1  # Remove this if a wc-like count is needed

    if time_print:
        end = time.time()
        print(round((end - start) * 1000, 2), 'ms')

    return count


def listdir(path, target='', exception=False, sep=False, logic_and=True):
    """
    An expanded function for os.listdir.
    :param path: get files and folders under this given path. (not necessary end with '/')
    :param target: element with such string will be selected (cooperated with param 'sep' and 'logic_and')
    :param exception: element with such string will not be selected (cooperated with param 'sep' and 'logic_and')
    :param sep: param 'target' and 'exception' will be split by sep
    :param logic_and: splited param 'target' and 'exception' will be used in the logic 'and', 'or'
    :return: file list after selcted
    """
    file_list = os.listdir(path)
    if sep:
        target = target.split(sep)
    else:
        target = [target]

    if logic_and:
        file_list = [i for i in file_list if (not sum([0 if tar in i else 1 for tar in target])) and ('.' != i[0])]
    else:
        file_list = [i for i in file_list if sum([1 if tar in i else 0 for tar in target]) and ('.' != i[0])]

    if exception:
        if sep:
            exception = exception.split(sep)
        else:
            exception = [exception]

        if logic_and:
            file_list = [i for i in file_list if not sum([1 if exc in i else 0 for exc in exception])]
        else:
            file_list = [i for i in file_list if sum([0 if exc in i else 1 for exc in exception])]

    return file_list


def reform_size(size_, detail=False, remain=3, rounding=2):
    if size_ < 0:
        raise ValueError('size can not be a negative number.')
    if size_ < 1024:
        return str(size_) + 'B'
    import math

    magnitude_name = ['B', 'KB', 'MB', 'GB', 'TB', 'PB', 'EB', 'ZB', 'YB']
    magnitude = int(math.log(size_, 1024))

    if not detail:
        size__ = np.round(size_ / (1024 ** (magnitude)), rounding)

        return str(size__) + magnitude_name[magnitude]

    else:
        size__ = size_
        infomation = []
        for i_ in range(magnitude):
            size__, size__0 = divmod(size__, 1024)
            infomation += [[size__, size__0]]

        out_num = infomation[-1]
        for info in infomation[:-1][::-1]:
            out_num += [info[1]]

        out_string_ = []
        for num_, name_ in zip(out_num, magnitude_name[: magnitude + 1][::-1]):
            if num_ != 0:
                out_string_ += [str(num_) + name_]

        return ' '.join(out_string_[: remain])


def getdirsize(path_, **kwargs):
    """
    Get whole size of the direction
    :param path_: direction path
    :param kwargs: keyword arguments for reform_size()
    :return: size
    """
    size_ = 0
    for root_, dirs_, files_ in os.walk(path_):
        size_ += sum([os.path.getsize(os.path.join(root_, file_)) for file_ in files_])

    return reform_size(size_, **kwargs)


# 给出路径下所有文件的文件夹, 快速
def get_all_file_fast(path_):
    """
    Get all file paths under the direction in a fast way
    :param path_: direction path
    :return: all file paths
    """
    all_path_ = []
    for root_, dirs_, files_ in os.walk(path_):
        all_path_ += [[os.path.join(root_, '').replace('\\', '/'), os.path.join(root_, file_).replace('\\', '/')] for
                      file_ in files_]

    return all_path_


# 给出路径下所有文件的文件夹, 包含了st.listdir效果
def get_all_file(path_, thres=99999, **kwargs):
    """
    Get all file paths under the direction in a detailed way, by kwargs used in list_selection().
    :param path_: direction path
    :param thres: file number limit to escape from unaffordable task. Default is 99999.
    :param kwargs: to select paths with keyword arguments used in list_selection().
    :return: all selected file paths
    """
    from .basic_tools import list_selection

    path_use = [path_]
    all_file_path = []
    while True:
        new_path_use = []
        for path_u in path_use:
            targets = listdir(path_u)
            for target in targets:
                if os.path.isfile(path_u + target):
                    all_file_path += [path_u + target]
                else:
                    new_path_use += [path_u + target + '/']

        if len(new_path_use) == 0:
            print('All files finished! NICE!')
            break
        elif len(all_file_path) > thres:
            print('Files more than thres! Use \'thres\' parameter to change the threshold number.')
            break
        path_use = new_path_use

    all_file_path = list_selection(all_file_path, **kwargs)

    return all_file_path


def path_diagnosis(path_):
    if path_[-1] != '/' or path_[-1] != '\\':
        path_ = path_ + '/'
        
    return path_


# pkl文件写入 读取
def read_pickle(path_):
    """
    Read a pickle file.
    :param path_: file path
    :return: obj
    """
    import pickle
    with open(path_, 'rb') as f:
        data_ = pickle.load(f)

    return data_


def write_pickle(path_, data_):
    """
    Write the object to a pickle file.
    :param path_: file save path
    :param data_: obj
    :return: None
    """
    import pickle
    with open(path_, 'wb+') as f:
        pickle.dump(data_, f)


def read_json(path_):
    """
    Read json file as a dictionary
    :param path_: json file path
    :return: dictionary obj
    """
    with open(path_, 'r') as file_:
        config_ = json.load(file_)

    return config_


def write_json(path_, json_file_):
    """
    Save the dictionary obj as a json file
    :param path_: json file save path
    :param json_file_: dictionary obj
    :return: None
    """
    with open(path_, 'w+') as file_:
        json.dump(json_file_, file_)


# excel 相关函数
if openpyxl_exist:
    def _read_sheet(sheets_, sheet_identify):
        if type(sheet_identify) == int:
            sheet1_ = sheets_[sheet_identify]
        elif type(sheet_identify) == str:
            sheet_use = np.argmax([i_.title == sheet_identify for i_ in sheets_])
            sheet1_ = sheets_[sheet_use]

        # 迭代读取所有的行
        rows = sheet1_.rows
        data_ = []
        for i_, row in enumerate(rows):
            row_val = [col.value for col in row]
            if i_ == 0:
                columns_use_ = row_val
            else:
                data_ += [row_val]
        data_ = pd.DataFrame(data_, columns=columns_use_)

        return data_


    def read_xlsx(path_, sheet_identify=0, dict_out=False):
        """
        Read excel file in a freedom framework
        :param path_: excel file path
        :param sheet_identify: int, str, list, or None, determine sheet(s) to be read
        :param dict_out: if out dict or list of dfs
        :return: df or list or dict of dfs in values
        """
        wb_ = load_workbook(path_)
        sheets = wb_.worksheets
        is_iter = isinstance(sheet_identify, Iterable)
        if not is_iter:
            return _read_sheet(sheets, sheet_identify)
        else:
            if not dict_out:
                return [_read_sheet(sheets, identify) for identify in sheet_identify]
            else:
                out_dict_ = {}
                for identify in sheet_identify:
                    out_dict_[identify] = _read_sheet(sheets, identify)

                return out_dict_


    def _input_sheet(ws_, df_):
        index_value = df_.index.names
        info = list(index_value) + list(df_.columns)
        ws_.append(info)

        for i_ in range(df_.shape[0]):
            index_value = df_.index[i_]
            is_iter = isinstance(index_value, Iterable)
            if not is_iter:
                index_value = [index_value]
            info = list(index_value) + list(df_.iloc[i_])
            ws_.append(info)

        return ws_


    def save_xlsx(save_path_, df_list_, titles=None):
        """
        Save df, list of dfs, or dict of dfs to excel file
        :param save_path_: excel file path
        :param df_list_: df, list of dfs, or dict of dfs
        :param titles: giving the sheet names of each dfs if df_list_'s type is list
        :return:
        """
        is_df = type(df_list_) == pd.core.frame.DataFrame
        if is_df:
            df_list_ = [df_list_]

        is_dict = type(df_list_) == dict
        if not is_dict:
            if not titles:
                titles = ['sheet' + str(i_) for i_ in range(1, len(df_list_) + 1)]
            out_dict_ = {}
            for title, df_ in zip(titles, df_list_):
                out_dict_[title] = df_

        wb_ = Workbook()
        for i_, zip_info in enumerate(zip(out_dict_.keys(), out_dict_.values())):
            ws_ = wb_.create_sheet(zip_info[0], i_)
            ws_ = _input_sheet(ws_, zip_info[1])

        wb_.save(save_path_)


# universal read framework
def read_file(path_, sep=False, sheet_identify=0, split=None, num_limit=0, decode='utf-8', **kwargs):
    """
    A universal read framework
    :param path_: file path
    :param sep: delimiter to use
    :param sheet_identify: str, int, list, or None
    :param split: fot fa and fq files, id info will split by the given separation and keep the front part
    :param num_limit: fot fa and fq files, stop reading at which line
    :param decode: fot fa and fq files, decode parameter
    :param kwargs: keyword argument for table like file
    :return: file
    """
    file_name = path_.rsplit('/', 1)[-1]
    houzhui_ = file_name.rsplit('.', 1)[-1]
    if houzhui_ in ['txt', 'tsv']:
        if not sep:
            sep = '\t'
        return pd.read_csv(path_, sep=sep, **kwargs)

    elif houzhui_ in ['csv']:
        if not sep:
            sep = ','
        return pd.read_csv(path_, sep=sep, **kwargs)

    elif houzhui_ in ['xls', 'xlsx', 'xlsm', 'xlsb', 'odf', 'ods', 'odt']:
        return pd.read_excel(path_, sheet_name=sheet_identify, **kwargs)

    elif houzhui_ in ['pkl']:
       return read_pickle(path_)

    elif houzhui_ in ['fa', 'fasta']:
        from .bio_tools import read_fasta
        return read_fasta(path_, split)

    elif '.fastq' in file_name or '.fq' in file_name:
        from .bio_tools import read_fastq, read_fastq_gz
        if houzhui_ in ['gz']:
            return read_fastq_gz(path_, split, num_limit, decode)
        else:
            return read_fastq(path_, split)

    elif houzhui_ in ['pdb']:
        from .bio_tools import read_pdb_file
        return read_pdb_file(path_)

    elif houzhui_ in ['json']:
        return read_json(path_)

    else:
        raise ValueError('File type .' + houzhui_ + ' has not been added to this function yet. ')


def _mark_code_path(path_, name_, replace=False):
    _path = path_ + 'result_from_code.txt'
    isExist = os.path.exists(_path)

    if replace:
        _info_dict = {}
        if isExist:
            with open(_path, 'r') as _file:
                for _line in _file:
                    _line_info = _line.strip('\n').split('\t')
                    _info_dict[_line_info[0]] = _line_info[1]
        _info_dict[name_] = _code_path

        with open(_path, 'w+') as _file:
            for _key, _value in zip(_info_dict.keys(), _info_dict.values()):
                _string = _key + '\t' + _value + '\n'
                _file.write(_string)

    else:
        if isExist:
            with open(_path, 'a+') as _file:
                _string = name_ + '\t' + _code_path + '\n'
                _file.write(_string)
        else:
            with open(_path, 'w+') as _file:
                _string = name_ + '\t' + _code_path + '\n'
                _file.write(_string)


# universal write framework
def write_file(path_, data_, sep=False, sheet_identify='sheet1', mark_code_path=True, replace=False, **kwargs):
    """
    A universal write framework
    :param path_: file path
    :param data_: data to save
    :param sep: delimiter to use
    :param sheet_identify: str, int, list, or None
    :param mark_code_path: mark where the code is within a config-like file
    :param replace: if replace the same result path in the config-like file
    :param kwargs: keyword argument for table like file
    :return: None
    """
    file_name = path_.rsplit('/', 1)[-1]
    houzhui_ = path_.rsplit('.', 1)[-1]
    if houzhui_ in ['txt', 'tsv']:
        if not sep:
            sep = '\t'
        data_.to_csv(path_, sep=sep, **kwargs)

    elif houzhui_ in ['csv']:
        if not sep:
            sep = ','
        data_.to_csv(path_, sep=sep, **kwargs)

    elif houzhui_ in ['xls', 'xlsx', 'xlsm', 'xlsb', 'odf', 'ods', 'odt']:
        data_.to_excel(path_, sheet_name=sheet_identify, **kwargs)

    elif houzhui_ in ['pkl']:
       write_pickle(path_, data_)

    elif houzhui_ in ['fa', 'fasta']:
        from .bio_tools import write_fasta
        write_fasta(path_, data_)

    elif 'fastq' in file_name or 'fq' in file_name:
        from .bio_tools import write_fastq, write_fastq_gz
        if houzhui_ in ['gz']:
            write_fastq_gz(path_, data_)
        else:
            write_fastq(path_, data_)

    elif houzhui_ in ['pdb']:
        from .bio_tools import read_pdb_file
        write_pickle(path_, data_)

    elif houzhui_ in ['json']:
        write_json(path_, data_)

    else:
        raise ValueError('File type .' + houzhui_ + ' has not been added to this function yet. ')

    if mark_code_path:
        _mark_code_path(path_.rsplit('/', 1)[0] + '/', path_.rsplit('/', 1)[-1], replace=replace)
